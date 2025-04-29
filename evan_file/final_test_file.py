from classical_algorithm.foronoi import Polygon
from algorithm.deterministic_algorithm import DeterministicAlgorithm
import csv
import gc
from openpyxl import load_workbook
import tracemalloc

k = 29

while k > 1:
    if k == 1:
        n = 3000
        # column_time = "BE"
        column_memory = "BF"
    if k == 2:
        n = 2500
        # column_time = "BC"
        column_memory = "BD"
    if k == 3:
        n = 2000
        # column_time = "BA"
        column_memory = "BB"
    if k == 4:
        n = 1500
        # column_time = "AY"
        column_memory = "AZ"
    if k == 5:
        n = 1400
        # column_time = "AW"
        column_memory = "AX"
    if k == 6:
        n = 1300
        # column_time = "AU"
        column_memory = "AV"
    if k == 7:
        n = 1200
        # column_time = "AS"
        column_memory = "AT"
    if k == 8:
        n = 1100
        # column_time = "AQ"
        column_memory = "AR"
    if k == 9:
        n = 1000
        # column_time = "AO"
        column_memory = "AP"
    if k == 10:
        n = 900
        # column_time = "AM"
        column_memory = "AN"
    if k == 11:
        n = 800
        # column_time = "AK"
        column_memory = "AL"
    if k == 12:
        n = 700
        # column_time = "AI"
        column_memory = "AJ"
    if k == 13:
        n = 600
        # column_time = "AG"
        column_memory = "AH"
    if k == 14:
        n = 500
        # column_time = "AE"
        column_memory = "AF"
    if k == 15:
        n = 400
        # column_time = "AC"
        column_memory = "AD"
    if k == 16:
        n = 300
        # column_time = "AA"
        column_memory = "AB"
    if k == 17:
        n = 250
        # column_time = "Y"
        column_memory = "Z"
    if k == 18:
        n = 200
        # column_time = "W"
        column_memory = "X"
    if k == 19:
        n = 150
        # column_time = "U"
        column_memory = "V"
    if k == 20:
        n = 100
        # column_time = "S"
        column_memory = "T"
    if k == 21:
        n = 90
        # column_time = "Q"
        column_memory = "R"
    if k == 22:
        n = 80
        # column_time = "O"
        column_memory = "P"
    if k == 23:
        n = 70
        # column_time = "M"
        column_memory = "N"
    if k == 24:
        n = 60
        # column_time = "K"
        column_memory = "L"
    if k == 25:
        n = 50
        # column_time = "I"
        column_memory = "J"
    if k == 26:
        n = 40
        # column_time = "G"
        column_memory = "H"
    if k == 27:
        n = 30
        # column_time = "E"
        column_memory = "F"
    if k == 28:
        n = 20
        # column_time = "C"
        column_memory = "D"
    if k == 29:
        n = 10
        # column_time = "A"
        column_memory = "B"

    # Define a set of points


    print(f"current n = {n}")

    seed = 1
    iteration_max = 10
    # batch_size = n - 1
    batch_size = n
    print(f"batch size = {batch_size}")

    while seed < iteration_max + 1:
        print(f"current seed = {seed}")
        # gc.disable()

        # Define a set of points
        points = []

        with open(f'INPUT/{n}points_seeds.csv', 'r', newline='') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if int(row['seed']) == seed:
                    x = float(row['x'])
                    y = float(row['y'])
                    points.append((x, y))

        # Define a bounding box
        polygon = Polygon([
            (-1000000, -1000000),
            (-1000000, 1000000),
            (1000000, 1000000),
            (1000000, -1000000)
        ])

        tracemalloc.start()

        # Initialize the algorithm
        v = DeterministicAlgorithm(polygon, batch_size)

        current_1, peak_1 = tracemalloc.get_traced_memory()

        # Create the diagram
        v.algorithm_process(points=points)

        current_2, peak_2 = tracemalloc.get_traced_memory()

        middle_memory = peak_2 - current_1
        print(f"middle_memory = {middle_memory / 1024} KB")

        gc.collect()

        current_3, peak_3 = tracemalloc.get_traced_memory()

        rubbish_memory = current_2 - current_3

        print(f"rubbish_memory = {rubbish_memory / 1024} KB")

        final_memory = middle_memory - rubbish_memory

        print(f"final_memory = {final_memory / 1024} KB")

        # gc.enable()
        tracemalloc.stop()

        excel_path = 'analysis_s_n_copy.xlsx'
        wb = load_workbook(excel_path)
        ws = wb.active
        row = 3 + seed
        # ws[f'{column_time}{row}'] = round(execution_time, 6)  # 秒
        ws[f'{column_memory}{row}'] = round(final_memory / 1024, 4)  # KB
        wb.save(excel_path)

        seed += 1
    k -= 1
print(f"Finished!")